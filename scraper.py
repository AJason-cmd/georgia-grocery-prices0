import asyncio
import re
import os
import requests
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

eggs = "\u10d9\u10d5\u10d4\u10e0\u10ea\u10ee\u10d8"
bread = "\u10de\u10e3\u10e0\u10d8"
rice = "\u10d1\u10e0\u10d8\u10dc\u10ef\u10d8"
oil = "\u10db\u10d6\u10d4\u10e1\u10e3\u10db\u10d6\u10d8\u10e0\u10d8\u10e1 \u10d6\u10d4\u10d7\u10d8"
butter = "\u10d9\u10d0\u10e0\u10d0\u10e5\u10d8"
sugar = "\u10e8\u10d0\u10e5\u10d0\u10e0\u10d8"
flour = "\u10e4\u10e5\u10d5\u10d8\u10da\u10d8"

ITEMS = [
    ("Eggs", eggs),
    ("Bread", bread),
    ("Rice", rice),
    ("Sunflower Oil", oil),
    ("Butter", butter),
    ("Sugar", sugar),
    ("Flour", flour),
]

STORES = [
    ("Spar", "https://sparonline.ge/search?q={q}", "https://sparonline.ge"),
    ("Ori Nabiji", "https://2nabiji.ge/en/search?query={q}", "https://2nabiji.ge"),
]

NAME_SELS = [
    "[class*='product-name']",
    "[class*='product-title']",
    "[class*='ProductName']",
    "[class*='name']",
    "[class*='title']",
    "h2",
    "h3",
    "h4",
]

PRICE_SELS = [
    "[class*='price']",
    "[class*='Price']",
    "[class*='cost']",
    "[class*='amount']",
    "span.price",
    "div.price",
]

IMG_SELS = [
    "img[class*='product']",
    "img[class*='Product']",
    "picture img",
    "img",
]

LINK_SELS = [
    "a[href*='/product']",
    "a[href*='/item']",
    "a[href*='/p/']",
    "a",
]

CARD_SELS = [
    ".product-card",
    ".product-item",
    "[class*='product-card']",
    "[class*='ProductCard']",
    "[class*='product_card']",
    "[class*='product-item']",
    "[class*='ProductItem']",
    ".item-box",
    "[class*='catalog-item']",
    "li.item",
    "article",
]


async def close_popups(page):
    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(1000)
    except Exception:
        pass
    close_sels = [
        "button[class*='close']",
        "button[class*='Close']",
        "[class*='modal'] button[class*='close']",
        "[class*='popup'] button",
        "button[aria-label='Close']",
        ".close-btn",
        ".btn-close",
        "[class*='dismiss']",
    ]
    for sel in close_sels:
        try:
            btn = await page.query_selector(sel)
            if btn and await btn.is_visible():
                await btn.click()
                await page.wait_for_timeout(1000)
                print("  Closed popup: " + sel)
                break
        except Exception:
            pass


async def get_text(el, sels):
    for s in sels:
        try:
            found = await el.query_selector(s)
            if found:
                t = (await found.inner_text()).strip()
                if t:
                    return t
        except Exception:
            pass
    return None


async def get_attr(el, sels, attr):
    for s in sels:
        try:
            found = await el.query_selector(s)
            if found:
                val = await found.get_attribute(attr)
                if val and len(val) > 3:
                    return val
        except Exception:
            pass
    return None


async def scrape(page, store_name, url_tpl, base_url, item_name, item_q):
    rows = []
    url = url_tpl.replace("{q}", item_q)
    os.makedirs("debug", exist_ok=True)

    try:
        print("  GET " + url)
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(6000)
        await close_popups(page)
        await page.wait_for_timeout(2000)
        shot_path = "debug/" + store_name + "_" + item_name + ".png"
        await page.screenshot(path=shot_path, full_page=False)

        for cs in CARD_SELS:
            cards = await page.query_selector_all(cs)
            if not cards or len(cards) < 2:
                continue
            print("  " + str(len(cards)) + " cards (" + cs + ")")
            for card in cards[:6]:
                name = await get_text(card, NAME_SELS)
                price_raw = await get_text(card, PRICE_SELS)
                img_src = await get_attr(card, IMG_SELS, "src")
                if not img_src:
                    img_src = await get_attr(card, IMG_SELS, "data-src")
                link_href = await get_attr(card, LINK_SELS, "href")

                if name and price_raw:
                    nums = re.findall(r'\d+[.,]?\d*', price_raw)
                    if nums:
                        price = float(nums[0].replace(',', '.'))
                        if 0.3 <= price <= 200:
                            if link_href and not link_href.startswith("http"):
                                link_href = base_url + link_href
                            if img_src and not img_src.startswith("http"):
                                img_src = base_url + img_src
                            rows.append({
                                "date": datetime.now().strftime("%Y-%m-%d"),
                                "store": store_name,
                                "category": item_name,
                                "product": name[:80],
                                "price": price,
                                "image": img_src or "",
                                "link": link_href or url,
                            })
            if rows:
                break

        print("  Result: " + str(len(rows)) + " prices")

    except Exception as e:
        print("  ERROR: " + str(e))

    return rows


def download_image(img_url):
    try:
        headers = {"User-Agent": "Mozilla/5.0 Chrome/124.0"}
        r = requests.get(img_url, timeout=8, headers=headers)
        ctype = r.headers.get("Content-Type", "")
        if r.status_code == 200 and "image" in ctype:
            return BytesIO(r.content)
    except Exception:
        pass
    return None


def build_excel(all_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bs = Side(style="thin", color="CCCCCC")
    border = Border(left=bs, right=bs, top=bs, bottom=bs)

    headers = ["Image", "Date", "Store", "Category", "Product", "Price (GEL)", "Link"]
    col_widths = [18, 12, 14, 14, 40, 14, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    all_rows.sort(key=lambda r: (r["category"], r["price"]))

    for row_num, r in enumerate(all_rows, start=2):
        ws.row_dimensions[row_num].height = 65

        img_data = download_image(r["image"]) if r["image"] else None
        if img_data:
            try:
                img = XLImage(img_data)
                img.width = 80
                img.height = 80
                img.anchor = "A" + str(row_num)
                ws.add_image(img)
            except Exception:
                ws.cell(row=row_num, column=1, value="(no image)")
        else:
            ws.cell(row=row_num, column=1, value="(no image)")

        vals = [None, r["date"], r["store"], r["category"], r["product"], r["price"]]
        for col, val in enumerate(vals, start=1):
            if val is None:
                continue
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = center
            cell.border = border
            if col == 6:
                cell.font = Font(bold=True, color="1F4E79", size=11)

        lc = ws.cell(row=row_num, column=7, value="Open product page")
        lc.hyperlink = r["link"]
        lc.font = Font(color="0563C1", underline="single")
        lc.alignment = center
        lc.border = border

        if row_num % 2 == 0:
            fill = PatternFill("solid", fgColor="EBF3FB")
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = fill

    wb.save("prices.xlsx")
    print("Saved prices.xlsx with " + str(len(all_rows)) + " rows")


async def main():
    all_rows = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
            ],
        )
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
        )
        page = await ctx.new_page()
        await page.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )

        for store_name, url_tpl, base_url in STORES:
            print("\n=== " + store_name + " ===")
            for item_name, item_q in ITEMS:
                print("  " + item_name + "...")
                rows = await scrape(page, store_name, url_tpl, base_url, item_name, item_q)
                all_rows.extend(rows)
                await asyncio.sleep(3)

        await browser.close()

    if all_rows:
        build_excel(all_rows)
        print("\nDone. " + str(len(all_rows)) + " prices saved to prices.xlsx")
    else:
        print("\nNo data collected - check debug screenshots")


asyncio.run(main())
